#spreadsheet library
from openpyxl import load_workbook

#GUI library
import PySimpleGUI as sg

#set layout of GUI window
layout = [[sg.T("")], [sg.Text("Select UCR request form:")], [sg.Input(key="file_path_input"), sg.FileBrowse()], 
         [sg.Text("Enter price:")], [sg.InputText(key="price")], 
         [sg.Text("Enter file destination folder:")], [sg.Input(key="-IN2-" ,change_submits=True), sg.FolderBrowse(key="destination_folder")],
         [sg.Button("Confirm")]]

#initialize the Window
window = sg.Window('CFN creator 1.0', layout, size = (425,300))

#default file_path, price, and destination_folder values
file_path = None
price = None
destination_folder = None

#opens window to get user input
while True:
    event, values = window.read()
    if event == sg.WIN_CLOSED or event=="Exit":
        break
    elif event == "Confirm":
        
        #gets file path to xlsx file
        file_path = values["file_path_input"]
        
        #throws error if invalid file type
        if !(file_path.endswith(".xlsx" || "xlsm" || "xltx" || "xltm")):
            raise Exception("Incorrect file type. Accepted file types include: xlsx/xlsm/xltx/xltm")
            
        #gets price from user input
        price = values["price"]
        
        #gets destination folder for CFN files form user input
        destination_folder = values["destination_folder"]
        
        
        
        window.close()

#loads spreadsheet from user-entered file path
wb = load_workbook(file_path)
sheet = wb.active

#getting column with UCR-IDs

#declaring default column value
UCR_ID_col = None

#declaring max column value for loop
max_col = sheet.max_column

for col in range(1, max_col + 1):
  cell = sheet.cell(row = 1, column = col)
  if cell.value == "UCR ID":
    UCR_ID_col = col
    break

#makes sure UCR-ID column was indentifed
if(UCR_ID_col == None):
    raise Exception("Could not find UCR-ID column; column should contain text \"UCR ID\""
    
#gets UCR-IDs from spreadsheet

#gets highest row value
num_rows = sheet.max_row   

for i in range(2, num_rows):
  
  #gets cell value containing UCR-ID
  UCR_ID = sheet.cell(row = i, column = UCR_ID_col).value
  
  #if column is empty, stop looping
  if UCR_ID == None:
    break
    
  #corrects UCR-ID naming for Windows file names
  file_name = UCR_ID.replace('/', '-')
  
  #creates CFN file 
  file = open(destination_folder + "\\" + file_name + ".CFN", 'x')
  file.write("SMARTECARTE_ENABLED=1\nSMARTECARTE_DEFAULT_AMOUNT="+price+"\nSMARTECARTE_UCR_ID="+UCR_ID)
  file.flush()
  file.close
  
exit()